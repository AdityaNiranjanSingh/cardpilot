from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO, StringIO
import csv
import re
from typing import Iterable

from openpyxl import load_workbook

from ..utils import canonical_category, parse_bool, parse_date, parse_decimal

HEADER_ALIASES = {
    "transaction_date": {
        "transaction_date",
        "date",
        "txn_date",
        "transaction date",
        "posting date",
        "post date",
        "transaction dt",
        "txn date",
        "value date",
    },
    "merchant_raw": {
        "merchant_raw",
        "merchant raw",
        "merchant",
        "description",
        "details",
        "transaction details",
        "merchant name",
        "particulars",
        "narration",
        "transaction description",
    },
    "amount_inr": {
        "amount_inr",
        "amount inr",
        "amount",
        "debit",
        "debit amount",
        "transaction amount",
        "spend",
        "inr amount",
        "purchase amount",
        "bill amount",
    },
    "actual_value_inr": {
        "actual_value_inr",
        "actual value inr",
        "actual reward",
        "cashback",
        "cashback credited",
        "reward value",
        "actual cashback",
        "rewards",
        "reward credited",
        "reward points value",
    },
    "transaction_channel": {"transaction_channel", "transaction channel", "channel", "mode", "online_offline", "online/offline"},
    "category": {"category", "merchant_category", "spend category", "spend type"},
    "mcc": {"mcc", "merchant category code"},
    "is_emi": {"is_emi", "is emi", "emi"},
    "is_wallet_load": {"is_wallet_load", "is wallet load", "wallet", "wallet_load", "wallet load"},
}

CSV_ENCODINGS = ("utf-8-sig", "utf-8", "cp1252", "latin-1")
MONTHS = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}
MONTH_PATTERN = "|".join(sorted(MONTHS, key=len, reverse=True))
DATE_PATTERNS = [
    re.compile(r"\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b"),
    re.compile(rf"\b\d{{1,2}}[\s-]+(?:{MONTH_PATTERN})\.?[,]?[\s-]*\d{{0,4}}\b", re.IGNORECASE),
    re.compile(rf"\b(?:{MONTH_PATTERN})\.?[\s-]+\d{{1,2}}[,]?[\s-]+\d{{2,4}}\b", re.IGNORECASE),
]
MONEY_RE = re.compile(
    r"(?P<currency>INR|Rs\.?|₹)?\s*(?P<number>[-+]?\d{1,3}(?:,\d{2,3})+(?:\.\d{1,2})?|[-+]?\d+(?:\.\d{1,2})?)\s*(?P<marker>Dr\.?|DR|Debit|Cr\.?|CR|Credit)?",
    re.IGNORECASE,
)
REWARD_RE = re.compile(r"(?:cashback|reward(?:s)?|points? credited|rp)\s*[:\-]?\s*(?:INR|Rs\.?|₹)?\s*(\d+(?:\.\d{1,2})?)", re.IGNORECASE)
LINE_SPLIT_BEFORE_DATE_RE = re.compile(
    r"(?=\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b|\b\d{1,2}[\s-]+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Sept|Oct|Nov|Dec)\b)",
    re.IGNORECASE,
)

SKIP_PDF_LINE_HINTS = {
    "minimum amount",
    "total amount",
    "total due",
    "amount due",
    "previous balance",
    "payment received",
    "payment thank",
    "thank you payment",
    "statement date",
    "payment due date",
    "credit limit",
    "available credit",
    "opening balance",
    "closing balance",
    "reward summary",
    "points summary",
    "cashback summary",
    "finance charges",
    "late payment",
    "tax invoice",
    "gstin",
    "summary of account",
    "domestic transactions",
    "international transactions",
    "page ",
}

MERCHANT_STOP_WORDS_RE = re.compile(
    r"\b(?:dr|cr|debit|credit|pos|ecom|upi|txn|transaction|auth|approval|ref(?:erence)?|number|no|card|ending|xx+|available|limit|balance)\b",
    re.IGNORECASE,
)


def _normalize_header(value: object) -> str:
    return str(value or "").strip().lower().replace("_", " ")


def _canonical_header(header: object) -> str | None:
    normalized = _normalize_header(header)
    for canonical, aliases in HEADER_ALIASES.items():
        if normalized in aliases:
            return canonical
    return None


def _decode_csv_content(content: bytes) -> str:
    last_error: Exception | None = None
    for encoding in CSV_ENCODINGS:
        try:
            return content.decode(encoding)
        except UnicodeDecodeError as exc:
            last_error = exc
    raise ValueError(f"Could not read CSV encoding. Please save the file as UTF-8 CSV or XLSX. Details: {last_error}")


def _clean_row(row: dict) -> dict | None:
    merchant = row.get("merchant_raw")
    amount = parse_decimal(row.get("amount_inr"))
    if not merchant or amount is None:
        return None
    if amount < 0:
        amount = abs(amount)
    category = row.get("category")
    return {
        "transaction_date": parse_date(row.get("transaction_date")) or date.today(),
        "merchant_raw": str(merchant).strip(),
        "amount_inr": amount,
        "actual_value_inr": parse_decimal(row.get("actual_value_inr")),
        "transaction_channel": row.get("transaction_channel") or "unknown",
        "category": category,
        "mcc": row.get("mcc"),
        "is_emi": bool(parse_bool(row.get("is_emi")) or False),
        "is_wallet_load": bool(parse_bool(row.get("is_wallet_load")) or False),
    }


def _csv_reader_for_text(text: str) -> csv.DictReader:
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    return csv.DictReader(StringIO(text), dialect=dialect)


def parse_csv_bytes(content: bytes) -> list[dict]:
    text = _decode_csv_content(content)
    reader = _csv_reader_for_text(text)
    mapped_headers = {h: _canonical_header(h) for h in (reader.fieldnames or [])}
    required_headers = set(mapped_headers.values())
    missing = {"merchant_raw", "amount_inr"} - required_headers
    if missing:
        known = ", ".join(reader.fieldnames or [])
        raise ValueError(
            "CSV is missing required columns. Required: merchant/description and amount/debit. "
            f"Found headers: {known}"
        )
    rows = []
    for raw in reader:
        mapped = {canonical: raw.get(original) for original, canonical in mapped_headers.items() if canonical}
        clean = _clean_row(mapped)
        if clean:
            rows.append(clean)
    return rows


def parse_xlsx_bytes(content: bytes) -> list[dict]:
    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Could not read XLSX file. Please upload a valid .xlsx file. Details: {exc}") from exc
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if not headers:
        return []
    canonical_headers = [_canonical_header(h) for h in headers]
    if "merchant_raw" not in canonical_headers or "amount_inr" not in canonical_headers:
        known = ", ".join(str(h or "") for h in headers)
        raise ValueError(
            "XLSX is missing required columns. Required: merchant/description and amount/debit. "
            f"Found headers: {known}"
        )
    rows = []
    for values in rows_iter:
        mapped = {canonical_headers[i]: values[i] for i in range(min(len(values), len(canonical_headers))) if canonical_headers[i]}
        clean = _clean_row(mapped)
        if clean:
            rows.append(clean)
    return rows


def _extract_statement_year(text: str) -> int:
    years = [int(y) for y in re.findall(r"\b20\d{2}\b", text)]
    if years:
        # The most common year in a statement is usually the statement/transaction year.
        return max(set(years), key=years.count)
    return date.today().year


def _parse_month_token(value: str) -> int | None:
    return MONTHS.get(value.lower().strip(" .,"))


def _parse_pdf_date(value: str, default_year: int | None = None) -> date | None:
    raw = " ".join(value.replace("-", " ").replace(",", " ").split())
    direct = parse_date(raw) or parse_date(value)
    if direct:
        return direct

    default_year = default_year or date.today().year
    numeric = re.match(r"^(\d{1,2})[/-](\d{1,2})(?:[/-](\d{2,4}))?$", value.strip())
    if numeric:
        day = int(numeric.group(1))
        month = int(numeric.group(2))
        year_text = numeric.group(3)
        year = int(year_text) if year_text else default_year
        if year < 100:
            year += 2000
        try:
            return date(year, month, day)
        except ValueError:
            return None

    day_month = re.match(rf"^(\d{{1,2}})\s+({MONTH_PATTERN})\.?\s*(\d{{2,4}})?$", raw, re.IGNORECASE)
    if day_month:
        day = int(day_month.group(1))
        month = _parse_month_token(day_month.group(2))
        year_text = day_month.group(3)
        year = int(year_text) if year_text else default_year
        if year < 100:
            year += 2000
        if month:
            try:
                return date(year, month, day)
            except ValueError:
                return None

    month_day = re.match(rf"^({MONTH_PATTERN})\.?\s+(\d{{1,2}})\s+(\d{{2,4}})$", raw, re.IGNORECASE)
    if month_day:
        month = _parse_month_token(month_day.group(1))
        day = int(month_day.group(2))
        year = int(month_day.group(3))
        if year < 100:
            year += 2000
        if month:
            try:
                return date(year, month, day)
            except ValueError:
                return None
    return None


def _find_first_date(line: str):
    best = None
    for pattern in DATE_PATTERNS:
        match = pattern.search(line)
        if match and (best is None or match.start() < best.start()):
            best = match
    return best


def _looks_like_credit_marker(text: str) -> bool:
    return bool(re.search(r"\b(cr|credit)\.?\b", text, re.IGNORECASE))


def _looks_like_debit_marker(text: str) -> bool:
    return bool(re.search(r"\b(dr|debit)\.?\b", text, re.IGNORECASE))


def _money_from_match(match: re.Match) -> Decimal | None:
    cleaned = (match.group("number") or "").replace(",", "").strip()
    if not cleaned:
        return None
    # Avoid card numbers, references, phone numbers, and years being treated as spend amounts.
    digits = re.sub(r"\D", "", cleaned)
    if len(digits) > 10:
        return None
    try:
        value = Decimal(cleaned)
    except Exception:
        return None
    if value <= 0:
        return None
    # Four-digit years without a currency/decimal/comma/marker are usually not transaction amounts.
    if 1900 <= int(value) <= 2099 and "." not in cleaned and "," not in (match.group(0) or "") and not (match.group("currency") or match.group("marker")):
        return None
    return value


def _choose_amount(matches: list[re.Match], line: str) -> tuple[Decimal | None, re.Match | None]:
    candidates: list[tuple[int, Decimal, re.Match]] = []
    for match in matches:
        value = _money_from_match(match)
        if value is None:
            continue
        marker = (match.group("marker") or "").lower()
        immediate_pre = line[max(0, match.start() - 6): match.start()]
        immediate_post = line[match.end(): match.end() + 8]
        immediate_context = f"{immediate_pre} {immediate_post}"
        if "cr" in marker or "credit" in marker or _looks_like_credit_marker(immediate_context):
            score = 20
        elif "dr" in marker or "debit" in marker or _looks_like_debit_marker(immediate_context):
            score = 110
        elif "." in match.group(0) or "," in match.group(0) or (match.group("currency") or ""):
            score = 75
        else:
            score = 40
        # Prefer the right-most amount when scores tie; amount usually appears near the end of the row.
        score += min(match.start() // 20, 8)
        candidates.append((score, value, match))
    if not candidates:
        return None, None
    candidates.sort(key=lambda item: (item[0], item[2].start()), reverse=True)
    best_score, best_value, best_match = candidates[0]
    if best_score < 30:
        return None, None
    return best_value, best_match


def _clean_pdf_merchant(value: str) -> str:
    merchant = value
    # Remove a second posting date if the PDF row includes txn date + posting date.
    first_date = _find_first_date(merchant)
    if first_date and first_date.start() <= 3:
        merchant = merchant[first_date.end():]
    merchant = re.sub(r"\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b", " ", merchant)
    merchant = re.sub(rf"\b\d{{1,2}}\s+(?:{MONTH_PATTERN})\.?\s*\d{{0,4}}\b", " ", merchant, flags=re.IGNORECASE)
    merchant = MERCHANT_STOP_WORDS_RE.sub(" ", merchant)
    merchant = re.sub(r"\b\d{5,}\b", " ", merchant)
    merchant = re.sub(r"[*#|:;]+", " ", merchant)
    merchant = re.sub(r"\s+", " ", merchant).strip(" -/")
    return merchant[:180]


def _normalize_pdf_line(line: str) -> str:
    line = line.replace("\u00a0", " ")
    line = line.replace("₹", " ₹")
    line = re.sub(r"\s+", " ", line).strip()
    return line


def _iter_pdf_candidate_lines(text: str) -> Iterable[str]:
    # First pass: existing text lines.
    for raw_line in text.splitlines():
        line = _normalize_pdf_line(raw_line)
        if line:
            yield line
    # Second pass: split long extracted blocks that contain many dated transactions on one line.
    compact = _normalize_pdf_line(text)
    for part in LINE_SPLIT_BEFORE_DATE_RE.split(compact):
        part = _normalize_pdf_line(part)
        if part:
            yield part


def _infer_channel_and_flags(merchant: str, line: str) -> tuple[str, bool, bool, str | None]:
    text = f"{merchant} {line}".lower()
    channel = "online" if any(token in text for token in ["online", "ecom", "amazon", "flipkart", "myntra", "swiggy", "zomato", "paytm", "upi"]) else "offline"
    if any(token in text for token in ["pos", "store", "retail"]):
        channel = "offline"
    is_emi = " emi" in f" {text} "
    is_wallet = any(token in text for token in ["wallet", "paytm wallet", "mobikwik", "amazon pay balance"])
    category = canonical_category(merchant)
    return channel, is_emi, is_wallet, category or None


def _actual_reward_from_line(line: str) -> Decimal | None:
    match = REWARD_RE.search(line)
    if not match:
        return None
    return parse_decimal(match.group(1))


def _row_from_pdf_line(line: str, default_year: int) -> dict | None:
    if not line or len(line) < 8:
        return None
    lowered = line.lower()
    if any(hint in lowered for hint in SKIP_PDF_LINE_HINTS):
        return None
    date_match = _find_first_date(line)
    if not date_match:
        return None
    txn_date = _parse_pdf_date(date_match.group(0), default_year=default_year)
    if not txn_date:
        return None

    after_date = line[date_match.end():].strip(" -:|*")
    # If the first token after the transaction date is another date/posting date, keep removing dates.
    for _ in range(2):
        second_date = _find_first_date(after_date)
        if second_date and second_date.start() <= 3:
            after_date = after_date[second_date.end():].strip(" -:|*")

    money_matches = list(MONEY_RE.finditer(after_date))
    if not money_matches:
        return None
    amount, amount_match = _choose_amount(money_matches, after_date)
    if amount is None or amount_match is None:
        return None
    if _looks_like_credit_marker(after_date[amount_match.start(): amount_match.end() + 10]) and not _looks_like_debit_marker(after_date):
        # Payments/refunds are credits and should not be treated as spends in the MVP analyzer.
        return None

    merchant = _clean_pdf_merchant(after_date[: amount_match.start()])
    if not merchant or len(merchant) < 3:
        return None
    if any(skip in merchant.lower() for skip in ["payment received", "payment thank", "autopay"]):
        return None

    channel, is_emi, is_wallet, category = _infer_channel_and_flags(merchant, line)
    return {
        "transaction_date": txn_date,
        "merchant_raw": merchant,
        "amount_inr": amount,
        "actual_value_inr": _actual_reward_from_line(line),
        "transaction_channel": channel,
        "category": category,
        "mcc": None,
        "is_emi": is_emi,
        "is_wallet_load": is_wallet,
    }


def _dedupe_rows(rows: list[dict]) -> list[dict]:
    seen: set[tuple] = set()
    unique: list[dict] = []
    for row in rows:
        merchant_key = re.sub(r"\W+", "", str(row.get("merchant_raw", "")).lower())[:60]
        key = (row.get("transaction_date"), merchant_key, str(row.get("amount_inr")))
        if key in seen:
            continue
        seen.add(key)
        unique.append(row)
    return unique


def parse_pdf_text(text: str) -> list[dict]:
    default_year = _extract_statement_year(text)
    rows = []
    for line in _iter_pdf_candidate_lines(text):
        row = _row_from_pdf_line(line, default_year=default_year)
        if row:
            rows.append(row)
    return _dedupe_rows(rows)


def _extract_with_pypdf(content: bytes, password: str | None) -> str:
    from pypdf import PdfReader

    reader = PdfReader(BytesIO(content))
    if reader.is_encrypted:
        if not password:
            raise ValueError("PDF is password-protected. Enter the statement password and upload again.")
        decrypt_result = reader.decrypt(password)
        if decrypt_result == 0:
            raise ValueError("Could not unlock PDF. Check the statement password.")

    page_texts: list[str] = []
    for page in reader.pages:
        try:
            layout_text = page.extract_text(extraction_mode="layout") or ""
        except TypeError:
            layout_text = page.extract_text() or ""
        except Exception:
            layout_text = page.extract_text() or ""
        page_texts.append(layout_text)
    return "\n".join(page_texts)


def _extract_with_pdfplumber(content: bytes, password: str | None) -> str:
    try:
        import pdfplumber
    except Exception:
        return ""
    chunks: list[str] = []
    with pdfplumber.open(BytesIO(content), password=password or "") as pdf:
        for page in pdf.pages:
            text = page.extract_text(x_tolerance=1, y_tolerance=3) or ""
            if text:
                chunks.append(text)
            try:
                for table in page.extract_tables() or []:
                    for row in table or []:
                        cells = [str(cell or "").strip() for cell in row or [] if str(cell or "").strip()]
                        if cells:
                            chunks.append(" | ".join(cells))
            except Exception:
                # Table extraction is helpful but optional.
                pass
    return "\n".join(chunks)


def parse_pdf_bytes(content: bytes, password: str | None = None) -> list[dict]:
    if not content:
        raise ValueError("Uploaded PDF is empty.")

    extraction_errors: list[str] = []
    text_candidates: list[str] = []

    try:
        text_candidates.append(_extract_with_pypdf(content, password=password))
    except ValueError:
        raise
    except Exception as exc:
        extraction_errors.append(f"pypdf: {exc}")

    try:
        plumber_text = _extract_with_pdfplumber(content, password=password)
        if plumber_text:
            text_candidates.append(plumber_text)
    except Exception as exc:
        extraction_errors.append(f"pdfplumber: {exc}")

    combined_text = "\n".join(text for text in text_candidates if text and text.strip())
    if len(combined_text.strip()) < 20:
        details = "; ".join(extraction_errors[:2])
        raise ValueError(
            "This PDF appears to be scanned/image-based or locked in a format that text extraction cannot read. "
            "Please upload the statement password if required, or export the statement as CSV/XLSX. "
            f"Details: {details}".strip()
        )

    rows = parse_pdf_text(combined_text)
    if not rows:
        sample = " | ".join(_normalize_pdf_line(line) for line in combined_text.splitlines()[:8] if _normalize_pdf_line(line))[:700]
        raise ValueError(
            "The PDF opened successfully, but CardPilot could not identify transaction rows in this bank format yet. "
            "Try a statement PDF where transaction rows contain date, merchant, and debit amount on the same row, or upload CSV/XLSX. "
            f"Detected text sample: {sample}"
        )
    return rows


def parse_statement_file(filename: str, content: bytes, password: str | None = None) -> list[dict]:
    lower = filename.lower()
    if lower.endswith(".csv"):
        return parse_csv_bytes(content)
    if lower.endswith(".xlsx") or lower.endswith(".xlsm"):
        return parse_xlsx_bytes(content)
    if lower.endswith(".xls"):
        raise ValueError("Old .xls files are not supported. Please save/export the file as .xlsx or .csv.")
    if lower.endswith(".pdf"):
        return parse_pdf_bytes(content, password=password)
    raise ValueError("Unsupported file type. Please upload CSV, XLSX, XLSM, or a text-based PDF statement.")
