from __future__ import annotations

from pathlib import Path
import sys

from openpyxl import load_workbook
from sqlalchemy.orm import Session

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.config import SEED_WORKBOOK_PATH
from app.db import SessionLocal, init_db
from app.models import Bank, Card, Merchant, RewardRule, RuleExclusion
from app.utils import parse_bool, parse_date, parse_decimal

SHEET_MODEL_MAP = {
    "Banks": Bank,
    "Cards": Card,
    "Reward_Rules": RewardRule,
    "Rule_Exclusions": RuleExclusion,
    "Merchants": Merchant,
}

PRIMARY_KEYS = {
    "Banks": "bank_id",
    "Cards": "card_id",
    "Reward_Rules": "rule_id",
    "Rule_Exclusions": "exclusion_id",
    "Merchants": "merchant_id",
}

MODEL_FIELD_ALIASES = {
    "Banks": {"bank_name": "bank_name"},
    "Cards": {"bank": "bank_name"},
}

DATE_FIELDS = {"last_verified", "valid_from", "valid_to"}
BOOL_FIELDS = {"needs_manual_review"}
NUMERIC_FIELDS = {"estimated_rate_pct", "estimated_reward_on_inr_1000", "parser_confidence", "current_year_spend_inr", "reward_balance"}

def normalize_row(sheet_name: str, headers: list[str], values: tuple) -> dict:
    row = {}
    aliases = MODEL_FIELD_ALIASES.get(sheet_name, {})
    for idx, header in enumerate(headers):
        if not header:
            continue
        key = str(header).strip()
        key = aliases.get(key, key)
        value = values[idx] if idx < len(values) else None
        if value == "":
            value = None
        if key in DATE_FIELDS:
            value = parse_date(value)
        elif key in BOOL_FIELDS:
            value = parse_bool(value)
        elif key in NUMERIC_FIELDS:
            value = parse_decimal(value)
        row[key] = value
    return row

def import_sheet(db: Session, wb, sheet_name: str) -> int:
    if sheet_name not in wb.sheetnames:
        return 0
    ws = wb[sheet_name]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    model = SHEET_MODEL_MAP[sheet_name]
    primary_key = PRIMARY_KEYS[sheet_name]
    model_columns = set(model.__mapper__.attrs.keys())
    count = 0
    for values in ws.iter_rows(min_row=2, values_only=True):
        raw_row = normalize_row(sheet_name, headers, values)
        if not raw_row.get(primary_key):
            continue
        row = {k: v for k, v in raw_row.items() if k in model_columns}
        existing = db.get(model, row[primary_key])
        if existing:
            for key, value in row.items():
                setattr(existing, key, value)
        else:
            db.add(model(**row))
        count += 1
    db.commit()
    return count

def import_workbook(path: str | Path = SEED_WORKBOOK_PATH) -> dict:
    init_db()
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    summary = {}
    with SessionLocal() as db:
        # Order matters because rules reference cards and exclusions reference rules.
        for sheet in ["Banks", "Cards", "Reward_Rules", "Rule_Exclusions", "Merchants"]:
            summary[sheet] = import_sheet(db, wb, sheet)
    return summary

if __name__ == "__main__":
    workbook_path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(SEED_WORKBOOK_PATH)
    summary = import_workbook(workbook_path)
    print("Imported workbook:", workbook_path)
    for sheet, count in summary.items():
        print(f"- {sheet}: {count} rows")
